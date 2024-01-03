from win32com import client
from unittest import mock
# Set max font family value to 100
p = mock.patch('openpyxl.styles.fonts.Font.family.max', new=100)
p.start()

from tqdm import tqdm
from tqdm import TqdmWarning
import openpyxl
from openpyxl import load_workbook
import shutil
import time
import os
import time
from datetime import date
import json
import sys

today = date.today()

class Autosave():
    def __init__(self,src,select,branch,mouth) -> None:
        self.branch = branch
        self.select = select
        self.mouth = mouth
        self.file = src
        self.savews = None
        self.temporaries = self.get_temporaries()
        self.sources = self.get_sources()["sources"]
        self.salib = self.get_sources()["salib"]
        self.temporary=self.get_sources()["temporary"]
        self.bar = None
        self.updatebar = None

    def calbar(self):
        percentage= 100
        target = 0
        if not self.select :
            sheets = []
            for sheet in self.sources:
                for i in range(self.get_round(sheet)):
                    target += 1
                    sheets.append(i)
            for i in sheets:
                target += 1
            result = (percentage/target)
            return result
        else:
            target = 2
            result = (percentage/target)
            return result

    def main(self):
        self.mkdir()
        self.updatebar = self.calbar()
        with tqdm(total=100, desc="Preparing", bar_format="{l_bar}{bar}|") as pbar:
            self.bar = pbar
            self.extract_convert(self.select)
        print("Successful")
        
    def get_temporaries(self):
        temporaries = 'temporary.xlsx'
        if os.path.exists(f"{temporaries}"):
            os.remove(f"{temporaries}")
        return temporaries

    def get_value(self,source,col,i):
        result = source.cell(row=i, column=col).value
        if result == None:
            result = "-"
        return result

    def get_round(self,source):
        i=0
        while source.cell(row=i+3, column=1).value != None:
            i+=1
        return i
    
    def get_sources(self):
        shutil.copyfile(self.file,self.temporaries)
        temporary = load_workbook(self.temporaries,data_only=True)
        salib = [i for i in temporary if "สลิป" in i.title and "Data" not in i.title]
        sources = [ i for i in temporary if "สลิป" not in i.title and "Data" not in i.title]
        return {
            "sources":sources,
            "salib":salib,
            "temporary":temporary
            }

    def mkdir(self):
        for i in self.sources:
            if i.title != "Data":
                path = f'{i.title}'
                if not os.path.exists(path):
                    os.makedirs(path)

    def get_lang(self,lang):
        try:
            with open(f"languages/{lang}.json", "r", encoding='utf-8') as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            return self.get_lang(self,'en')
        return data

    def extract_convert(self,person=False):
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        if not person:
            for sheet in self.sources:
                for i in self.temporary.sheetnames:
                    if i != "สลิป":
                        self.temporary.remove(self.temporary[i])
                file = []
                for i in range(self.get_round(sheet)):
                        i += 3
                        respound = self.get_lang(self.get_value(sheet,29,i))
                        img = openpyxl.drawing.image.Image('Harislogo.jpg')
                        img.anchor = 'B1'
                        ws = [i.title for i in self.salib]
                        salib = self.temporary[ws[0]]
                        salib.add_image(img)

                        salib["C1"] = respound["address"][sheet.title]["adline1"]
                        salib["C2"] = respound["address"][sheet.title]["adline2"]
                        salib["C3"] = respound["address"][sheet.title]["adline3"]
                        salib["B4"] = respound["branch"]
                        salib["B5"] = respound["personnelcode"]
                        salib["B6"] = respound["name"]
                        salib["B7"] = respound["position"]
                        salib["B9"] = respound["earnings"]
                        salib["B11"] = respound["salary"]
                        salib["B12"] = respound["positionallowance"]
                        salib["B13"] = respound["otd"]
                        salib["B14"] = respound["oth"]
                        salib["B15"] = respound["diligenceallowance"]
                        salib["B16"] = respound["welfare"]
                        salib["B17"] = respound["incentive"]
                        salib["B18"] = respound["bonus"]
                        salib["B20"] = respound["totale"]
                        salib["B22"] = respound["net"]
                        salib["A25"] = respound["warning"]
                        salib["F20"] = respound["totled"]
                        salib["F17"] = respound["loan"]
                        salib["F16"] = respound["debt"]
                        salib["F15"] = respound["leave"]
                        salib["F14"] = respound["late"]
                        salib["F13"] = respound["repayment"]
                        salib["F12"] = respound["social"]
                        salib["F11"] = respound["advance"]
                        salib["F9"] = respound["deduction"]
                        salib["J1"] = respound["payslip"]
                        salib["J2"] = respound["paymentdate"]
                        salib["K9"] = respound["details"]
                        salib["K11"] = respound["absent"]
                        salib["K12"] = respound["late"]
                        salib["K13"] = respound["sick"]
                        salib["K14"] = respound["personal"]
                        salib["K15"] = respound["vacation"]
                        salib["K16"] = respound["otd"]
                        salib["K17"] = respound["oth"]


                        salib["C4"] = respound[sheet.title] #สาขา
                        salib["C5"] = self.get_value(sheet,1,i) #รหัสพนักงาน
                        salib["C6"] = self.get_value(sheet,2,i) #ชื่อ-สกุล
                        salib["C7"] = self.get_value(sheet,4,i) #ตำเเหน่ง
                        salib["B8"] = f"{self.mouth}"
                        salib["C11"] = self.get_value(sheet,5,i) #อัตราเงินเดือน
                        salib["C12"] = self.get_value(sheet,6,i) #ค่าตำแหน่ง
                        salib["C13"] = self.get_value(sheet,7,i) #OT
                        salib["C14"] = self.get_value(sheet,8,i) #ค่าล่วงเวลา
                        salib["C15"] = self.get_value(sheet,9,i) #เบี้ยขยัน
                        salib["C16"] = self.get_value(sheet,14,i) #สวัสดิการอื่นๆ
                        salib["C17"] = self.get_value(sheet,18,i) #ยอดเป้า
                        salib["C18"] = self.get_value(sheet,19,i) #โบนัส
                        salib["G11"] = self.get_value(sheet,10,i) #เบิก
                        salib["G12"] = self.get_value(sheet,11,i) #ประกันสังคม
                        salib["G13"] = self.get_value(sheet,12,i) #ยอดจ่ายเงินกู้
                        salib["G14"] = self.get_value(sheet,16,i) #สาย
                        salib["G15"] = self.get_value(sheet,17,i) #ลา
                        salib["G16"] = self.get_value(sheet,15,i) #หนี้
                        salib["G17"] = self.get_value(sheet,13,i) #ยอดเงินกู้คงเหลือ
                        salib["L11"] = self.get_value(sheet,22,i) #ขาด(วัน)
                        salib["L12"] = self.get_value(sheet,23,i) #สาย(วัน)
                        salib["L13"] = self.get_value(sheet,24,i) #ลาป่วย(นาที)
                        salib["L14"] = self.get_value(sheet,25,i) #ลากิจ(วัน)
                        salib["L15"] = self.get_value(sheet,26,i) #ลาพักร้อน(วัน)
                        salib["L16"] = self.get_value(sheet,27,i) #OT(วัน)
                        salib["L17"] = self.get_value(sheet,28,i) #ล่วงเวลา(ชั่วโมง)
                        salib["C20"] = '=SUM(C11:C18)' #รวมเงินได้
                        salib["G20"] = '=SUM(G11:G16)' #รวมรายการหัก
                        salib["C22"] = self.get_value(sheet,20,i) #รายได้สุทธิ
                        salib["L2"] = today.strftime("%d/%m/%Y")
                        filename = f"{self.get_value(sheet,2,i)},{self.get_value(sheet,21,i)}"
                        self.temporary.save(f"{sheet.title}\\{filename}.xlsx")
                        file.append(salib['C6'].value)
                        self.bar.update(self.updatebar)
                        wb = app.Workbooks.Open(f"{os.getcwd()}\\{sheet.title}\\{filename}.xlsx")
                        wb.ActiveSheet.ExportAsFixedFormat(0,f"{os.getcwd()}\\{sheet.title}\\{filename}")
                        wb.Close()   
                        os.remove(f"{os.getcwd()}\{sheet.title}\{filename}.xlsx")
                        self.bar.desc = f"Extracting to PDF branch {sheet.title}"
                        self.bar.update(self.updatebar)
                shutil.copyfile(self.file,self.temporaries)
                self.temporary = load_workbook(self.temporaries,data_only=True)
        else:
            shutil.copyfile(self.file,self.temporaries)
            self.temporary = load_workbook(self.temporaries,data_only=True)
            sheet = self.temporary[self.branch]
            for i in self.temporary.sheetnames:
                    if i != "สลิป":
                        self.temporary.remove(self.temporary[i])
            print(self.temporary.sheetnames)
            salib = self.temporary["สลิป"]
            img = openpyxl.drawing.image.Image(f'{os.getcwd()}\\Harislogo.jpg')
            img.anchor = 'B1'
            salib.add_image(img)
            index = int(self.select)+2
            respound = self.get_lang(self.get_value(sheet,29,index))
            salib["C1"] = respound["address"][sheet.title]["adline1"]
            salib["C2"] = respound["address"][sheet.title]["adline2"]
            salib["C3"] = respound["address"][sheet.title]["adline3"]
            salib["B4"] = respound["branch"]
            salib["B5"] = respound["personnelcode"]
            salib["B6"] = respound["name"]
            salib["B7"] = respound["position"]
            salib["B9"] = respound["earnings"]
            salib["B11"] = respound["salary"]
            salib["B12"] = respound["positionallowance"]
            salib["B13"] = respound["otd"]
            salib["B14"] = respound["oth"]
            salib["B15"] = respound["diligenceallowance"]
            salib["B16"] = respound["welfare"]
            salib["B17"] = respound["incentive"]
            salib["B18"] = respound["bonus"]
            salib["B20"] = respound["totale"]
            salib["B22"] = respound["net"]
            salib["A25"] = respound["warning"]
            salib["F20"] = respound["totled"]
            salib["F17"] = respound["loan"]
            salib["F16"] = respound["debt"]
            salib["F15"] = respound["leave"]
            salib["F14"] = respound["late"]
            salib["F13"] = respound["repayment"]
            salib["F12"] = respound["social"]
            salib["F11"] = respound["advance"]
            salib["F9"] = respound["deduction"]
            salib["J1"] = respound["payslip"]
            salib["J2"] = respound["paymentdate"]
            salib["K9"] = respound["details"]
            salib["K11"] = respound["absent"]
            salib["K12"] = respound["late"]
            salib["K13"] = respound["sick"]
            salib["K14"] = respound["personal"]
            salib["K15"] = respound["vacation"]
            salib["K16"] = respound["otd"]
            salib["K17"] = respound["oth"]


            salib["C4"] = respound[sheet.title] #สาขา
            salib["C5"] = self.get_value(sheet,1,index) #รหัสพนักงาน
            salib["C6"] = self.get_value(sheet,2,index) #ชื่อ-สกุล
            salib["C7"] = self.get_value(sheet,4,index) #ตำเเหน่ง
            salib["B8"] = f"{self.mouth}"
            salib["C11"] = self.get_value(sheet,5,index) #อัตราเงินเดือน
            salib["C12"] = self.get_value(sheet,6,index) #ค่าตำแหน่ง
            salib["C13"] = self.get_value(sheet,7,index) #OT
            salib["C14"] = self.get_value(sheet,8,index) #ค่าล่วงเวลา
            salib["C15"] = self.get_value(sheet,9,index) #เบี้ยขยัน
            salib["C16"] = self.get_value(sheet,14,index) #สวัสดิการอื่นๆ
            salib["C17"] = self.get_value(sheet,18,index) #ยอดเป้า
            salib["C18"] = self.get_value(sheet,19,index) #โบนัส
            salib["G11"] = self.get_value(sheet,10,index) #เบิก
            salib["G12"] = self.get_value(sheet,11,index) #ประกันสังคม
            salib["G13"] = self.get_value(sheet,12,index) #ยอดจ่ายเงินกู้
            salib["G14"] = self.get_value(sheet,16,index) #สาย
            salib["G15"] = self.get_value(sheet,17,index) #ลา
            salib["G16"] = self.get_value(sheet,15,index) #หนี้
            salib["G17"] = self.get_value(sheet,13,index) #ยอดเงินกู้คงเหลือ
            salib["L11"] = self.get_value(sheet,22,index) #ขาด(วัน)
            salib["L12"] = self.get_value(sheet,23,index) #สาย(วัน)
            salib["L13"] = self.get_value(sheet,24,index) #ลาป่วย(นาที)
            salib["L14"] = self.get_value(sheet,25,index) #ลากิจ(วัน)
            salib["L15"] = self.get_value(sheet,26,index) #ลาพักร้อน(วัน)
            salib["L16"] = self.get_value(sheet,27,index) #OT(วัน)
            salib["L17"] = self.get_value(sheet,28,index) #ล่วงเวลา(ชั่วโมง)
            salib["C20"] = '=SUM(C11:C18)' #รวมเงินได้
            salib["G20"] = '=SUM(G11:G16)' #รวมรายการหัก
            salib["C22"] = self.get_value(sheet,20,index) #รายได้สุทธิ
            salib["L2"] = today.strftime("%d/%m/%Y")
            self.bar.desc = f"Extracting to Excel branch {sheet.title}"
            filename = f"{self.get_value(sheet,2,index)},{self.get_value(sheet,21,index)}"
            self.temporary.save(f"{sheet.title}\\{filename}.xlsx")
            self.bar.update(self.updatebar)
            wb = app.Workbooks.Open(f"{os.getcwd()}\\{sheet.title}\\{filename}.xlsx")
            wb.ActiveSheet.ExportAsFixedFormat(0,f"{os.getcwd()}\\{sheet.title}\\{filename}")
            wb.Close()
            os.remove(f"{os.getcwd()}\\{sheet.title}\\{filename}.xlsx")
        self.bar.desc = f"Extraction complete {sheet.title}"
        self.bar.close
        os.remove(self.temporaries)



