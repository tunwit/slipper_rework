from win32com import client
from unittest import mock
# Set max font family value to 100
p = mock.patch('openpyxl.styles.fonts.Font.family.max', new=100)
p.start()

import os
import json
import time
import time
from datetime import date
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import shutil

SHOP_NAME = 'haris'
creating = False

month = {
        "January":"มกราคม", 
        "February":"กุมภาพันธ์", 
        "March":"มีนาคม", 
        "April":"เมษายน", 
        "May":"พฤษภาคม", 
        "June":"มิถุนายน", 
        "July":"กรกฎาคม", 
        "August":"สิงหาคม", 
        "September":"กันยายน",
        "October":"ตุลาคม", 
        "November":"พฤศจิกายน", 
        "December":"ธันวาคม"
    }


class excel():
    def __init__(self,path) -> None:
        self.path = path
        self.output_dir = self.get_output_dir()
        self.temporaries = self.get_temporaries()
        self.sources = self.get_sources()["sources"]
        self.salib = self.get_sources()["salib"]
        self.temporary=self.get_sources()["temporary"]
        self.people = None
        self.call_back = None
        self.complete = False

    def get_output_dir(self):
        output = f"{os.getcwd()}\slip\{SHOP_NAME}"
        return output

    def re_init(self):
        self.people = None
        self.complete = False

    def get_temporaries(self):
        temporaries = 'temporary.xlsx'
        if os.path.exists(f"{temporaries}"):
            os.remove(f"{temporaries}")
        return temporaries
    
    def get_sources(self):
        shutil.copyfile(self.path,self.temporaries)
        temporary = load_workbook(self.temporaries,data_only=True)
        salib = [i for i in temporary if "สลิป" in i.title and "Data" not in i.title]
        sources = [ i for i in temporary if i.title.startswith('D') and "Data" not in i.title]
        return {
            "sources":sources,
            "salib":salib,
            "temporary":temporary
            }
    
    def get_lang(self,lang):
        try:
            with open(f"data\\languages\\{lang}.json", "r", encoding='utf-8') as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            with open(f"data\\languages\\th.json", "r", encoding='utf-8') as json_file:
                data = json.load(json_file)
        return data
    
    def get_value(self,source,col,i):
        result = source.cell(row=i, column=col).value
        if result == None:
            result = "-"
        return result
    
    def get_round(self,source):
        i=0
        while source.cell(row=i+3, column=1).value != None:
            i+=1
        return i
    
    def get_all_round(self):
        i=0
        for sheet in self.sources:
            salib = self.get_round(sheet)
            i += salib
        return i

    def mkdir(self,branch):
            if not os.path.exists(os.path.join(self.output_dir,branch)):
                os.makedirs(os.path.join(self.output_dir,branch))

    def progress(self,index=None,current=None,branch=None):
        if self.call_back:
            data = {
                'status':'complete'if self.complete else 'processing',
                'all':len(self.people),
            }
            if current:
                data.update({'current':current})
            if index:
                data.update({'index':index,
                             'percentage':(index*100)/len(self.people)})
            if index:
                data.update({'branch':branch})
            self.call_back(data)

    def extract_convert(self,people_pre:list,date_m:date):
        global creating
        creating = True
        people= []
        for person in people_pre:
            person:str
            new = person.replace('[/size]','')
            people.append(new)
        self.re_init()
        self.people = people
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        index = 0
        for sheet in self.sources:
            for i in self.temporary.sheetnames:
                if i != "สลิป":
                    self.temporary.remove(self.temporary[i])
            file = []
            for num,i in enumerate(range(self.get_round(sheet)),1):
                    sheet_title = sheet.title.replace('D','')
                    i += 3
                    if not self.get_value(sheet,2,i) in people:
                        continue
                    index += 1
                    self.mkdir(sheet_title)
                    self.progress(index,self.get_value(sheet,2,i),sheet_title)
                    respound = self.get_lang(self.get_value(sheet,27,i))
                    img = openpyxl.drawing.image.Image('data\\image\\Harislogo.jpg')
                    img.anchor = 'B1'
                    ws = [i.title for i in self.salib]
                    salib = self.temporary[ws[0]]
                    salib.add_image(img)

                    salib["C1"] = respound["address"][sheet_title]["adline1"]
                    salib["C2"] = respound["address"][sheet_title]["adline2"]
                    salib["C3"] = respound["address"][sheet_title]["adline3"]
                    salib["B4"] = respound["branch"]
                    salib["B5"] = respound["personnelcode"]
                    salib["B6"] = respound["name"]
                    salib["B7"] = respound["position"]
                    salib["B9"] = respound["earnings"]
                    salib["B11"] = respound["salary"]
                    salib["B12"] = respound["positionallowance"]
                    salib["B13"] = respound["otd"]
                    salib["B14"] = respound["oth"]
                    salib["B15"] = respound["otho"]
                    salib["B16"] = respound["diligenceallowance"]
                    salib["B17"] = respound["welfare"]
                    salib["B18"] = respound["target"]
                    salib["B19"] = respound["bonus"]
                    salib["B21"] = respound["totale"]
                    salib["B23"] = respound["net"]
                    salib["A25"] = respound["warning"]
                    salib["F21"] = respound["totled"]
                    salib["F17"] = respound["loan"]
                    salib["F16"] = respound["debt"]
                    salib["F15"] = respound["leave"]
                    salib["F14"] = respound["late"]
                    salib["F13"] = respound["repayment"]
                    salib["F12"] = respound["social"]
                    salib["F11"] = respound["advance"]
                    salib["F9"] = respound["deduction"]
                    salib["J1"] = respound["payslip"]
                    salib["K9"] = respound["details"]
                    salib["K11"] = respound["absent"]
                    salib["K12"] = respound["late"]
                    salib["K13"] = respound["sick"]
                    salib["K14"] = respound["personal"]
                    salib["K15"] = respound["vacation"]
                    salib["K16"] = respound["otd"]
                    salib["K17"] = respound["oth"]
                    salib["K18"] = respound["otho"]

                    salib["C4"] = sheet_title #สาขา
                    salib["C5"] = self.get_value(sheet,1,i) #รหัสพนักงาน
                    salib["C6"] = self.get_value(sheet,2,i) #ชื่อ-สกุล
                    salib["C7"] = self.get_value(sheet,4,i) #ตำเเหน่ง
                    # salib["B8"] = f"{respound['ofmonth']} {mouth[(date - relativedelta(months=1)).strftime('%B')]} {date.year}"
                    salib["B8"] = f"{respound['ofmonth'].format(month=month[date_m.strftime('%B')],year=date_m.year)}"
                    salib["C11"] = self.get_value(sheet,5,i) #อัตราเงินเดือน
                    salib["C12"] = self.get_value(sheet,6,i) #ค่าตำแหน่ง
                    salib["C13"] = self.get_value(sheet,7,i) #OT (วัน)
                    salib["C14"] = self.get_value(sheet,8,i) #OT (ชั่วโมง)
                    salib["C15"] = self.get_value(sheet,9,i) #OT (ชั่วโมงวันขัตฤกษ์)
                    salib["C16"] = self.get_value(sheet,10,i) #เบี้ยขยัน
                    salib["C17"] = self.get_value(sheet,15,i) #สวัสดิการอื่นๆ
                    salib["C18"] = self.get_value(sheet,19,i) #ยอดเป้า
                    salib["C19"] = self.get_value(sheet,20, i) #โบนัส
                    salib["G11"] = self.get_value(sheet,11,i) #เบิก
                    salib["G12"] = self.get_value(sheet,12,i) #ประกันสังคม
                    salib["G13"] = self.get_value(sheet,13,i) #ยอดจ่ายเงินกู้
                    salib["G14"] = self.get_value(sheet,17,i) #สาย
                    salib["G15"] = self.get_value(sheet,18,i) #ลา
                    salib["G16"] = self.get_value(sheet,16,i) #หนี้
                    salib["G17"] = self.get_value(sheet,14,i) #ยอดเงินกู้คงเหลือ
                    salib["L11"] = self.get_value(sheet,23,i) #ขาด
                    salib["L12"] = self.get_value(sheet,24,i) #สาย (นาที)
                    salib["L13"] = self.get_value(sheet,25,i) #ลาป่วย
                    salib["L14"] = self.get_value(sheet,26,i) #ลากิจ(วัน)
                    salib["L15"] = self.get_value(sheet,27,i) #ลาพักร้อน(วัน)
                    salib["L16"] = self.get_value(sheet,28,i) #OT (วัน)
                    salib["L17"] = self.get_value(sheet,29,i) #OT (ชั่วโมง)
                    salib["L18"] = self.get_value(sheet,30,i) #OT (ชั่วโมงวันขัตฤกษ์)
                    salib["C21"] = '=SUM(C11:C19)' #รวมเงินได้
                    salib["G21"] = '=SUM(G11:G16)' #รวมรายการหัก
                    salib["C23"] = self.get_value(sheet,21,i) #รายได้สุทธิ
                    filename = f"{self.get_value(sheet,2,i)},{self.get_value(sheet,22,i)},0,{date_m.strftime('%B')},{datetime.now().strftime('%d%m%y%H%M%S')}"
                    finalpath_ex = os.path.join(self.output_dir,sheet_title,f"{filename}.xlsx")
                    self.temporary.save(finalpath_ex)
                    file.append(salib['C6'].value)
                    time.sleep(0.3)
                    wb = app.Workbooks.Open(finalpath_ex)
                    wb.ActiveSheet.PageSetup.Orientation = 2
                    wb.ActiveSheet.PageSetup.Zoom = False
                    wb.ActiveSheet.PageSetup.FitToPagesTall = 1
                    wb.ActiveSheet.PageSetup.FitToPagesWide = 1
                    wb.ActiveSheet.ExportAsFixedFormat(0,os.path.join(self.output_dir,sheet_title,f"{filename}"))
                    wb.Save()
                    wb.Close()   
                    time.sleep(0.1)
                    os.remove(finalpath_ex)
            shutil.copyfile(self.path,self.temporaries)
            self.temporary = load_workbook(self.temporaries,data_only=True)
        os.remove(self.temporaries)
        self.complete = True
        self.progress()
        creating = False