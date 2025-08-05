from unittest import mock
# Set max font family value to 100
p = mock.patch('openpyxl.styles.fonts.Font.family.max', new=100)
p.start()

import os
import json
from pathlib import Path
import time
from datetime import date
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import shutil
from setup_config import SLIP_DETAIL, LOGO_PATH, SHOP_NAME
from openpyxl.styles import Font
import pandas as pd
from pathlib import Path
from jinja2 import Environment, FileSystemLoader
import asyncio
from pyppeteer import launch
import threading

creating = False

month = {
        "January":"มกราคม", 
        "February":"กุมภาพันธ์", 
        "March":"มีนาคม", 
        "April":"เมษายน", 
        "May":"พฤษภาคม", 
        "June":"มิถุนายน", 
        "July":"กรกฎาคม", 
        "August":"สิงหาคม", 
        "September":"กันยายน",
        "October":"ตุลาคม", 
        "November":"พฤศจิกายน", 
        "December":"ธันวาคม"
    }


class excel():
    def __init__(self,path) -> None:
        self.path = path
        self.storage_dir = self.get_storage_dir()
        self.slip_dir = self.get_slip_dir()
        self.dfs = self.ex_to_df()
        self.call_back = None
        self.complete = False
        self.total_rows = 0
        self.jobs = []
        self.progress_percent = 0

    def ex_to_df(self):
        dfs = pd.read_excel(self.path,header=1,sheet_name=None)
        return self.clean_sheet(dfs)
    
    def clean_sheet(self,dfs:dict):
        to_pop = [df for df in dfs if not df.startswith("D")]
        for b in to_pop:
            dfs.pop(b)

        cleaned_dfs = {}
        for sheet_name, df in dfs.items():
            df.columns.values[0] = "รหัสพนักงาน"
    
            if "รหัสพนักงาน" in df.columns and "ชื่อ-นามสกุล" in df.columns:
                new_name = sheet_name.replace("D", "", 1)  # remove only the first "D"
                cleaned_dfs[new_name] = df.dropna(subset=["รหัสพนักงาน", "ชื่อ-นามสกุล"])
                first_col = cleaned_dfs[new_name].columns[0]
                cleaned_dfs[new_name][first_col] = cleaned_dfs[new_name][first_col].astype(pd.Int64Dtype()).astype(str)
                cleaned_dfs[new_name].fillna(0, inplace=True)

        print(cleaned_dfs)
        return cleaned_dfs
    
    def get_storage_dir(self):
        output = Path.cwd() / "storage" / SHOP_NAME

        return output
    
    def get_slip_dir(self):
        output = Path.cwd() / "slip" / SHOP_NAME
        return output
    
    def re_init(self):
        self.total_rows = 0
        self.complete = False

    def get_lang(self,lang):
        path = Path.cwd() / "data" / "languages"
        try:
            with open(path / f"{lang}.json", "r", encoding='utf-8') as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            with open(path / "th.json", "r", encoding='utf-8') as json_file:
                data = json.load(json_file)
        return data
    
    def get_value(self,employee,col:str):
        result = employee[col]
        if result == None or result == '0':
            result = "-"
        return result

    def get_round(self,branch):
        return self.dfs[branch].shape[0]
    

    def mkdir(self,branch):
            if not os.path.exists(self.slip_dir / branch):
                os.makedirs(self.slip_dir / branch)
            if not os.path.exists(self.storage_dir / branch):
                os.makedirs(self.storage_dir / branch)

    def progress(self,current=None,branch=None,file='html'):
        self.progress_percent+=1
        if self.call_back:
            data = {
                'status':'complete'if self.complete else 'processing',
                'file':file,
                'all':self.total_rows*2,
            }
            if current:
                data.update({'current':current})
            print(f'{self.progress_percent} |{self.total_rows*2}')
            data.update({'percentage':(self.progress_percent*100)/(self.total_rows*2)})
            if branch:
                data.update({'branch':branch})
            self.call_back(data)


    def format_value(self,value, fmt):
        if fmt == "float":
            return f"{value:,.2f}"
        elif fmt == "int":
            return f"{int(value)}"
        else:
            return f"{value}"
    
    async def render_to_pdf(self,semaphore,browser, html_path, pdf_path, path_storage, context):
        async with semaphore:
                page = await browser.newPage()
                await page.goto(f'file:///{html_path}', waitUntil='networkidle0')
                await page.pdf({
                    'path': pdf_path,
                    'format': 'A4',
                    'landscape': True,
                    'printBackground': True,
                    'preferCSSPageSize': True,
                    'margin': {'top': '10mm', 'bottom': '10mm', 'left': '10mm', 'right': '10mm'},
                })
                #create pay slip pdf
                file_name = path_storage / "pay_slip"
                self.createMetaData(path_storage,context,file_name)
                key = f'{context['employee']["id"]}_{context['employee']['name']}'
                path_slip = self.slip_dir / context['employee']['branch'] / key
                shutil.copy2(file_name.with_suffix(".pdf"), path_slip.with_suffix(".pdf"))
                self.progress(current=context['employee']["name"],branch=context['employee']['branch'],file='pdf')

    async def html_to_pdf_batch(self):
            if os.name == 'nt':
                exe = Path("C:\Program Files\Google\Chrome\Application\chrome.exe")
            else:
                exe = Path(r"/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
            browser = await launch(headless=True,
                        executablePath=exe,handleSIGINT=False,
                        handleSIGTERM=False,
                        handleSIGHUP=False,
                        dumpio=False, 
                        args=['--no-sandbox'])
            semaphore = asyncio.Semaphore(5)
            tasks = []
            for html_path, pdf_path, path_storage, context in self.jobs:
                task = self.render_to_pdf(semaphore,browser, html_path, pdf_path, path_storage, context)
                tasks.append(task)
            # for html_path, pdf_path, path_strorage, context in self.jobs:
            #     await page.goto(f'file:///{html_path}', waitUntil='networkidle0')
            #     await page.pdf({
            #         'path': pdf_path,
            #         'format': 'A4',
            #         'landscape': True,
            #         'printBackground': True,
            #         'preferCSSPageSize': True,
            #         'margin': {'top': '10mm', 'bottom': '10mm', 'left': '10mm', 'right': '10mm'},
            #     })
            #     #create pay slip pdf
            #     file_name = path_strorage / "pay_slip"
            #     self.createMetaData(path_strorage,context,file_name)
            #     key = f'{context['employee']["id"]}_{context['employee']['name']}'
            #     path_slip = self.slip_dir / context['employee']['branch'] / key
            #     shutil.copy2(file_name.with_suffix(".pdf"), path_slip.with_suffix(".pdf"))
            await asyncio.gather(*tasks)
            await browser.close()


    def createMetaData(self,storage_path,context,file_name):
        meta_data = {
            "employee_id": context['employee']['id'],
            "employee_name": context['employee']['name'],
            "email": context['employee']['email'],
            "pay_period": context['payPeriod'],
            "pdf_path": str(file_name.with_suffix(".pdf")),
            "html_path": str(file_name.with_suffix(".html")),
            "mail_sent": False,
            "created_at": context['timeStamp']
        }
        json_file = storage_path / "metadata.json"
        with open(json_file,'w',encoding="utf-8") as f:
            json.dump(meta_data, f, ensure_ascii=False, indent=2)
        

    def build_section(self,employee,employee_data,config,date_m,t):
        earnings = [{"label" : t(field['label_key']),
                             "value":employee_data[field['key']],
                             "unit": t(field["unit_key"]),
                             "formatted": self.format_value(employee_data[field['key']], field['format_key'])}
                             for field in config['earnings']['fields'] if field['display']]
                
        deduction = [{"label" : t(field['label_key']),
                        "value":employee_data[field['key']],
                        "unit": t(field["unit_key"]),
                        "formatted": self.format_value(employee_data[field['key']], field['format_key'])} 
                        for field in config['deduction']['fields'] if field['display']]
        
        details = [{"label" : t(field['label_key']),
                        "value":employee_data[field['key']],
                        "unit": t(field["unit_key"]),
                        "formatted": self.format_value(employee_data[field['key']], field['format_key'])} 
                        for field in config['details']['fields'] if field['display']]
        
        net = {"label":t(config['total']['label_key']),
                "value":employee_data[config['total']['key']],
                "unit": t(config['total']["unit_key"]),
                "formatted": self.format_value(employee_data[config['total']['key']], config['total']['format_key'])}

        total_earnings = {
            "label":t('totale'),
            "value":sum(item["value"] for item in earnings),
            "unit": t('฿'),
            "formatted": self.format_value(sum(item["value"] for item in earnings),"float")}
        

        total_deduction = {
            "label":t('totled'),
            "value":sum(item["value"] for item in deduction),
            "unit": t('฿'),
            "formatted": self.format_value(sum(item["value"] for item in deduction),"float")
        }

        context = {
        "company":config['company'],
        "employee": employee,
        "sections": {
            "earnings": earnings,
            "deduction": deduction,
            "details":details,
            "net":net
        },
        "totals": {
                "earnings": total_earnings,
            "deduction": total_deduction,
        },
        "payPeriod":date_m.strftime("%B %Y"),
        "timeStamp":datetime.now().strftime("%d %B %Y %H:%M:%S")
        }
        return context
    
    def extract_convert(self,data:dict,date_m:date):
        global creating
        creating = True

        self.re_init()
        self.total_rows = sum(len(df) for df in data.values())
        env = Environment(loader=FileSystemLoader('data/template'))
        with open("config_2.json","r",encoding="utf8") as config:
            config = json.load(config)['haris_slip_details']

        template = env.get_template(config['template'])

        #pre load langauge
        lang_th = self.get_lang('th')
        lang_en = self.get_lang('en')

        for branch in data.keys():
            self.mkdir(branch)
            df_map = {row["รหัสพนักงาน"]: row for _, row in self.dfs[branch].iterrows()}
            for _id in data[branch]:
                employee_data = df_map[_id]

                if employee_data['ภาษา'] == 'th':
                    lang = lang_th
                else:
                    lang = lang_en

                def t(key):
                    return lang.get(key, key)
                
                employee = {
                    "name" : employee_data["ชื่อ-นามสกุล"],
                    "id": employee_data["รหัสพนักงาน"],
                    "position": employee_data["ตำแหน่ง"],
                    "email":employee_data['Email'],
                    "branch":branch
                }
                
                context = self.build_section(employee,employee_data,config,date_m,t)
               
                html_out = template.render(context=context,t=t)
                key = f'{employee["id"]}_{employee['name']}'
                path_strorage = self.storage_dir / branch / key 

                #create pay slip html
                os.makedirs(path_strorage, exist_ok=True)
                file_name = path_strorage / "pay_slip"
                with open(file_name.with_suffix(".html"),'w', encoding='utf-8') as f:
                    f.write(html_out)
                
                self.jobs.append((file_name.with_suffix(".html"), file_name.with_suffix(".pdf"),path_strorage,context))
                self.progress(current=employee["name"],branch=branch)

        asyncio.run(self.html_to_pdf_batch())
        self.complete = True
        self.progress()
        creating = False
        self.progress_percent = 0

        # for df in self.dfs:
        #     for i in self.temporary.sheetnames:
        #         if i != "สลิป":
        #             self.temporary.remove(self.temporary[i])
        #     for i in range(self.get_round(sheet)):
        #             sheet_title = sheet.title.replace('D','')
        #             i += 3
        #             if not self.get_value(sheet,2,i) in people:
        #                 continue
        #             index += 1
        #             self.mkdir(sheet_title)
        #             self.progress(index,self.get_value(sheet,2,i),sheet_title)
        #             respound = self.get_lang(self.get_value(sheet,27,i))
        #             img = openpyxl.drawing.image.Image(LOGO_PATH)
        #             img.anchor = 'B1'
        #             ws = [i.title for i in self.salib]
        #             salib = self.temporary[ws[0]]
        #             salib.add_image(img)
                    
        #             salib["C1"] = respound["address"][sheet_title]["adline1"]
        #             salib["C2"] = respound["address"][sheet_title]["adline2"]
        #             salib["C3"] = respound["address"][sheet_title]["adline3"]
        #             salib["C4"] = sheet_title #สาขา
        #             salib["B8"] = f"{respound['ofmonth'].format(month=month[date_m.strftime('%B')],year=date_m.year)}"
        #             # "Key":[text_pos,value_pos,col_pos]

        #             email = self.get_value(sheet,SLIP_DETAIL['email_col'],i)
        #             for field in SLIP_DETAIL.items():
        #                 if field[0] == "email_col":
        #                     continue
        #                 key,text_pos ,value_pos ,col_pos = field[0],field[1][0],field[1][1],field[1][2]
        #                 salib[text_pos] = respound[key]
        #                 if value_pos:
        #                     if type(col_pos) == int:
        #                         salib[value_pos] = self.get_value(sheet,col_pos,i)
        #                     else:
        #                         salib[value_pos] = col_pos
                    
        #             filename = f"{self.get_value(sheet,2,i)},{email},0,{date_m.strftime('%B')},{datetime.now().strftime('%d%m%y%H%M%S')}"
        #             print(filename)
        #             finalpath_ex = os.path.join(self.output_dir,sheet_title,f"{filename}.xlsx")
        #             self.temporary.save(finalpath_ex)
        #             # wb = app.Workbooks.Open(finalpath_ex)
        #             # wb.ActiveSheet.PageSetup.Orientation = 2
        #             # wb.ActiveSheet.PageSetup.Zoom = False
        #             # wb.ActiveSheet.PageSetup.FitToPagesTall = 1
        #             # wb.ActiveSheet.PageSetup.FitToPagesWide = 1
        #             # wb.ActiveSheet.ExportAsFixedFormat(0,os.path.join(self.output_dir,sheet_title,f"{filename}"))
        #             # wb.Save()
        #             # wb.Close()   
        #             os.remove(finalpath_ex)
        #     shutil.copyfile(self.path,self.temporaries)
        #     self.temporary = load_workbook(self.temporaries,data_only=True)
        #     time.sleep(0.1)
        # os.remove(self.temporaries)
        # self.complete = True
        # self.progress()
        # creating = False