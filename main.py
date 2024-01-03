from win32com import client
from unittest import mock
# Set max font family value to 100
p = mock.patch('openpyxl.styles.fonts.Font.family.max', new=100)
p.start()
from kivy.core.text import LabelBase
from kivymd.uix.button import MDRectangleFlatButton
from kivymd.uix.label import MDLabel
from kivymd.uix.boxlayout import BoxLayout
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.list import IconRightWidget
from kivymd.uix.list import IconLeftWidget
from kivymd.app import MDApp
from kivy.lang import Builder
from kivymd.uix.selectioncontrol import MDCheckbox 
from kivymd.uix.card import MDSeparator 
from kivymd.uix.scrollview import MDScrollView 
from kivymd.uix.list import MDList ,IRightBodyTouch
from kivymd.uix.button import MDRectangleFlatButton 
from kivymd.uix.dialog import MDDialog
from kivymd.uix.label import MDIcon
from kivymd.uix.button import MDFlatButton
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.snackbar import MDSnackbar
from kivymd.uix.progressbar import MDProgressBar
from kivymd.uix.textfield import MDTextField
from kivymd.uix.pickers import MDDatePicker
from kivy.clock import Clock
from kivymd.theming import ThemeManager
import os
import tkinter
from tkinter import filedialog
from kivymd.uix.list import OneLineAvatarIconListItem
from kivymd.font_definitions import theme_font_styles
import os
from kivy.properties import  StringProperty,BooleanProperty,ListProperty
from kivymd.uix.expansionpanel import MDExpansionPanel, MDExpansionPanelTwoLine
from tqdm import tqdm
from tqdm import TqdmWarning
import openpyxl
from openpyxl import load_workbook
import shutil
from io import BytesIO
import os
import time
from datetime import date
from datetime import datetime
import json
import sys
import threading
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import ssl
import smtplib
import json
from email.utils import formatdate
from email.mime.application import MIMEApplication
import pypdfium2 as pdfium
import tempfile
import re
from dateutil.relativedelta import relativedelta

creating = False
mouth = {
        "January":"มกราคม", 
        "February":"กุมภาพันธ์", 
        "March":"มีนาคม", 
        "April":"เมษายน", 
        "May":"พฤษภาคม", 
        "June":"มิถุนายน", 
        "July":"กรกฎาคม", 
        "August":"สิงหาคม", 
        "September":"กันยายน",
        "October":"ตุลาคม", 
        "November":"พฤศจิกายน", 
        "December":"ธันวาคม"
    }
KV = """

MDScreen: 
    MDBoxLayout:   
        MDNavigationRail:
            width: "100dp"
            current_selected_item: 0
            md_bg_color: 0.217,0.217,0.217,0.1
            anchor: "center"
            selected_color_background: "#e7e4c0"
            text_color_item_active: app.theme_cls.primary_color
            icon_color_item_active: app.theme_cls.primary_light
            MDNavigationRailItem:
                icon: "newspaper-variant-multiple-outline"
                text: "Slip Maker"
                on_press : mana.current = "SlipMaker"
            MDNavigationRailItem:
                icon: "gmail"
                text: "Gmail Sender"
                on_press : mana.current = "GmailSender"
            MDNavigationRailItem:
                icon: "account"
                text: "Employee"
                on_press : mana.current = "Employee"
            MDNavigationRailItem:
                icon: "cog"
                text: "Setting"
                on_press : mana.current = "Setting"

        MDScreenManager:
            id: mana
            SlipMaker:
            GmailSender:
            Employee:
            Setting:
        
<SlipMaker>:
    name: "SlipMaker"
    MDBoxLayout:
        id: box1
        orientation:"horizontal"
        MDScreen:
            MDLabel:
                text: root.excel_path
                halign: "center"
                font_name: "sarabunBold"
                font_size: "50sp"
                pos_hint:{"center_y":.6}
            MDRectangleFlatButton
                markup:True
                text: "[b]Browse File[/b]"
                font_name: "sarabun"
                font_size: "20sp"
                pos_hint:{"center_x":.5,"center_y":.5}
                on_press: root.on_browse()
        
<GmailSender>:
    name : "GmailSender"
    MDBoxLayout:
        orientation: "vertical"
        padding: '20dp'
        spacing: '5dp'
        MDBoxLayout:
            id: box_email
            size_hint:1,1
            orientation: "vertical"
            MDLabel:
                id: no_employee_label_email
                text: 'No employee deploy'
                font_name: "sarabun"
                font_size: "20sp"
                theme_text_color:'Hint'
                halign: 'center'

        MDBoxLayout:
            spacing: '20dp'
            size_hint:1,0.15
            orientation: "vertical"
            MDSeparator:
            MDFillRoundFlatIconButton:
                id: sendemail_button
                icon: 'email'
                font_name: "sarabun"
                font_size: "20sp"
                text: f' 0       Send email               '
                pos_hint:{'center_y':0.5,'center_x':0.5}
                halign: 'center'
                disabled:True
                on_press: root.send_email_button()

<Employee>:
    name : "Employee"
    MDBoxLayout:
        orientation: "horizontal"
        MDBoxLayout:
            id:box_employee
            orientation: "vertical"
            MDLabel:
                id: no_employee_label
                text: 'No employee deploy'
                font_name: "sarabun"
                font_size: "20sp"
                halign: 'center'
                pos_hint:{"center_y":.5}
                theme_text_color:'Hint'
                        

        MDBoxLayout:
            md_bg_color: (0.217,0.217,0.217,0.1)
            padding: "20dp"
            spacing: "10dp"
            orientation: "vertical"
            size_hint:0.3,1
            MDScreen:
                MDIcon:
                    icon:'account'
                    pos_hint:{"center_x":.04,"center_y":0.051}
                MDRaisedButton:
                    id: delete_button
                    text: "Delete"
                    font_name: "sarabun"
                    font_size: "20sp"
                    pos_hint:{"center_x":.7,"center_y":0.05}
                    disabled:True
                    on_press: root.delete_selected()
                MDLabel:
                    id: delete_label_count
                    text: "0"
                    font_name: "sarabun"
                    font_size: "20sp"
                    pos_hint:{"center_x":.65,"center_y":0.05}

<Setting>:
    name : "Setting"
    MDBoxLayout:
        orientation: "vertical"
        padding: "40dp"
        spacing: "10dp"
        MDBoxLayout:
            orientation: "horizontal"
            MDLabel:
                text:'Langauge'
                pos_hint:{'center_y':0.5}
            MDSwitch:
                halign:
        MDSeparator:


"""


class send_email():
    mouth = {
        "January":"มกราคม", 
        "February":"กุมภาพันธ์", 
        "March":"มีนาคม", 
        "April":"เมษายน", 
        "May":"พฤษภาคม", 
        "June":"มิถุนายน", 
        "July":"กรกฎาคม", 
        "August":"สิงหาคม", 
        "September":"กันยายน",
        "October":"ตุลาคม", 
        "November":"พฤศจิกายน", 
        "December":"ธันวาคม"
    }
    def __init__(self) -> None:
        self.people = None
        self.call_back = None
        self.complete = False
    
    def progress(self,index=None,current=None,branch=None):
        if self.call_back:
            data = {
                'status':'complete'if self.complete else 'processing',
                'all':len(self.people),
            }
            if current:
                data.update({'current':current})
            if index:
                data.update({'index':index,
                             'percentage':(index*100)/len(self.people)})
            if index:
                data.update({'branch':branch})
            self.call_back(data)

    def send(self,people):
        self.people = people
        with open("config.json","r",encoding="utf8") as config:
            data = json.load(config)
            
        sender = data["sender_email"]
        password = data["email_password"]
        context = ssl.create_default_context()

        # text = """
        # เรียน {name},\n\n
        # \tใบสลิปเงินเดือนของ {name} ประจำเดือน {DY} ของบริษัท ไอแอมฟู้ด จำกัด\n
        # หากมีข้อผิดพลาดประการใดขออภัยไว้ ณ ที่นี้\n\n
        # \t\t\t\t\tจึงเรียนมาเพื่อทราบ\n
        # \t\t\t\t\tบริษัท ไอแอมฟู้ด จำกัด
        # """
        with smtplib.SMTP_SSL('smtp.gmail.com',465,context = context) as smtp:
            smtp.login(sender,password) 
            index = 0
            for person in self.people:
                index += 1
                self.progress(index,person['name'],person['branch'])
                if person['email'] == "-":
                    continue
                msg = MIMEMultipart()
                msg['From'] = "Haris premium buffet"
                msg['To'] = person['email']
                msg['subject'] = f'สลิปเงินเดือนของ {person["name"]}'
                msg['Date'] = formatdate(localtime=True)

                # msg.attach(MIMEText(text.format(name=person.split(",")[0],DY=f"{mouth[time.strftime('%B')]} {int(time.strftime('%Y'))+543}")))

                msg.attach(MIMEText('<img src="cid:image1" width="1000" height="772">', 'html'))
                
                pdf = pdfium.PdfDocument(person['path'])
                page = pdf.get_page(0)
                pil_image = page.render(scale = 300/72).to_pil()
                pdf.close()

                temp_file = tempfile.NamedTemporaryFile(delete=False,suffix='.png')
                pil_image.save(temp_file.name)
                image_data = temp_file.read()
                temp_file.close()

                img = MIMEImage(image_data,_subtype="png")
                img.add_header('Content-ID', '<image1>')
                msg.attach(img)

                with open(person["path"],'rb') as f:   
                    attach = MIMEApplication(f.read(),_subtype="pdf")
                attach.add_header('Content-Disposition','attachment',filename=f"เงินเดือนของ {person['name']} ประจำเดือน {mouth[time.strftime('%B')]} {int(time.strftime('%Y'))+543}.pdf")
                msg.attach(attach)
                
                smtp.sendmail(sender,person['email'],msg.as_string())
                print(f'Mail has send to {person["name"]} | {person["email"]} {index}/{len(person)}')

                path_name:str = os.path.split(person["path"])
                splited_name = person['file_name'].split(',')
                splited_name[-1] = '1.pdf'
                new_name = ','.join(splited_name)
                new_path = os.path.join(path_name[0],new_name)
                try:
                    os.rename(person["path"],new_path)
                except:pass
            self.complete = True
            self.progress()
class excel():
    def __init__(self,path) -> None:
        self.path = path
        self.output_dir = r'C:\Users\Tunwi\Desktop\Python_project\Harisslip\slip'
        self.temporaries = self.get_temporaries()
        self.sources = self.get_sources()["sources"]
        self.salib = self.get_sources()["salib"]
        self.temporary=self.get_sources()["temporary"]
        self.people = None
        self.call_back = None
        self.complete = False

    def re_init(self):
        self.people = None
        self.complete = False

    def get_temporaries(self):
        temporaries = 'temporary.xlsx'
        if os.path.exists(f"{temporaries}"):
            os.remove(f"{temporaries}")
        return temporaries
    
    def get_sources(self):
        shutil.copyfile(self.path,self.temporaries)
        temporary = load_workbook(self.temporaries,data_only=True)
        salib = [i for i in temporary if "สลิป" in i.title and "Data" not in i.title]
        sources = [ i for i in temporary if "สลิป" not in i.title and "Data" not in i.title]
        return {
            "sources":sources,
            "salib":salib,
            "temporary":temporary
            }
    
    def get_lang(self,lang):
        try:
            with open(f"data\languages\{lang}.json", "r", encoding='utf-8') as json_file:
                data = json.load(json_file)
        except FileNotFoundError:
            return self.get_lang(self,'en')
        return data
    
    def get_value(self,source,col,i):
        result = source.cell(row=i, column=col).value
        if result == None:
            result = "-"
        return result
    
    def get_round(self,source):
        i=0
        while source.cell(row=i+3, column=1).value != None:
            i+=1
        return i
    
    def get_all_round(self):
        i=0
        for sheet in self.sources:
            salib = self.get_round(sheet)
            i += salib
        return i

    def mkdir(self,branch):
            if not os.path.exists(os.path.join(self.output_dir,branch)):
                os.makedirs(os.path.join(self.output_dir,branch))

    def progress(self,index=None,current=None,branch=None):
        if self.call_back:
            data = {
                'status':'complete'if self.complete else 'processing',
                'all':len(self.people),
            }
            if current:
                data.update({'current':current})
            if index:
                data.update({'index':index,
                             'percentage':(index*100)/len(self.people)})
            if index:
                data.update({'branch':branch})
            self.call_back(data)

    def extract_convert(self,people_pre:list,date_m:date):
        global creating
        creating = True
        people= []
        for person in people_pre:
            person:str
            new = person.replace('[/size]','')
            people.append(new)
        self.re_init()
        self.people = people
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        index = 0
        for sheet in self.sources:
            for i in self.temporary.sheetnames:
                if i != "สลิป":
                    self.temporary.remove(self.temporary[i])
            file = []
            for num,i in enumerate(range(self.get_round(sheet)),1):
                    i += 3
                    if not self.get_value(sheet,2,i) in people:
                        continue
                    index += 1
                    self.mkdir(sheet.title)
                    self.progress(index,self.get_value(sheet,2,i),sheet.title)
                    respound = self.get_lang(self.get_value(sheet,29,i))
                    img = openpyxl.drawing.image.Image('data\image\Harislogo.jpg')
                    img.anchor = 'B1'
                    ws = [i.title for i in self.salib]
                    salib = self.temporary[ws[0]]
                    salib.add_image(img)

                    salib["C1"] = respound["address"][sheet.title]["adline1"]
                    salib["C2"] = respound["address"][sheet.title]["adline2"]
                    salib["C3"] = respound["address"][sheet.title]["adline3"]
                    salib["B4"] = respound["branch"]
                    salib["B5"] = respound["personnelcode"]
                    salib["B6"] = respound["name"]
                    salib["B7"] = respound["position"]
                    salib["B9"] = respound["earnings"]
                    salib["B11"] = respound["salary"]
                    salib["B12"] = respound["positionallowance"]
                    salib["B13"] = respound["otd"]
                    salib["B14"] = respound["oth"]
                    salib["B15"] = respound["diligenceallowance"]
                    salib["B16"] = respound["welfare"]
                    salib["B17"] = respound["incentive"]
                    salib["B18"] = respound["bonus"]
                    salib["B20"] = respound["totale"]
                    salib["B22"] = respound["net"]
                    salib["A25"] = respound["warning"]
                    salib["F20"] = respound["totled"]
                    salib["F17"] = respound["loan"]
                    salib["F16"] = respound["debt"]
                    salib["F15"] = respound["leave"]
                    salib["F14"] = respound["late"]
                    salib["F13"] = respound["repayment"]
                    salib["F12"] = respound["social"]
                    salib["F11"] = respound["advance"]
                    salib["F9"] = respound["deduction"]
                    salib["J1"] = respound["payslip"]
                    salib["K9"] = respound["details"]
                    salib["K11"] = respound["absent"]
                    salib["K12"] = respound["late"]
                    salib["K13"] = respound["sick"]
                    salib["K14"] = respound["personal"]
                    salib["K15"] = respound["vacation"]
                    salib["K16"] = respound["otd"]
                    salib["K17"] = respound["oth"]

                    salib["C4"] = respound[sheet.title] #สาขา
                    salib["C5"] = self.get_value(sheet,1,i) #รหัสพนักงาน
                    salib["C6"] = self.get_value(sheet,2,i) #ชื่อ-สกุล
                    salib["C7"] = self.get_value(sheet,4,i) #ตำเเหน่ง
                    # salib["B8"] = f"{respound['ofmonth']} {mouth[(date - relativedelta(months=1)).strftime('%B')]} {date.year}"
                    salib["B8"] = f"{respound['ofmonth'].format(month=mouth[date_m.strftime('%B')],year=date_m.year)}"
                    salib["C11"] = self.get_value(sheet,5,i) #อัตราเงินเดือน
                    salib["C12"] = self.get_value(sheet,6,i) #ค่าตำแหน่ง
                    salib["C13"] = self.get_value(sheet,7,i) #OT
                    salib["C14"] = self.get_value(sheet,8,i) #ค่าล่วงเวลา
                    salib["C15"] = self.get_value(sheet,9,i) #เบี้ยขยัน
                    salib["C16"] = self.get_value(sheet,14,i) #สวัสดิการอื่นๆ
                    salib["C17"] = self.get_value(sheet,18,i) #ยอดเป้า
                    salib["C18"] = self.get_value(sheet,19,i) #โบนัส
                    salib["G11"] = self.get_value(sheet,10,i) #เบิก
                    salib["G12"] = self.get_value(sheet,11,i) #ประกันสังคม
                    salib["G13"] = self.get_value(sheet,12,i) #ยอดจ่ายเงินกู้
                    salib["G14"] = self.get_value(sheet,16,i) #สาย
                    salib["G15"] = self.get_value(sheet,17,i) #ลา
                    salib["G16"] = self.get_value(sheet,15,i) #หนี้
                    salib["G17"] = self.get_value(sheet,13,i) #ยอดเงินกู้คงเหลือ
                    salib["L11"] = self.get_value(sheet,22,i) #ขาด(วัน)
                    salib["L12"] = self.get_value(sheet,23,i) #สาย(วัน)
                    salib["L13"] = self.get_value(sheet,24,i) #ลาป่วย(นาที)
                    salib["L14"] = self.get_value(sheet,25,i) #ลากิจ(วัน)
                    salib["L15"] = self.get_value(sheet,26,i) #ลาพักร้อน(วัน)
                    salib["L16"] = self.get_value(sheet,27,i) #OT(วัน)
                    salib["L17"] = self.get_value(sheet,28,i) #ล่วงเวลา(ชั่วโมง)
                    salib["C20"] = '=SUM(C11:C18)' #รวมเงินได้
                    salib["G20"] = '=SUM(G11:G16)' #รวมรายการหัก
                    salib["C22"] = self.get_value(sheet,20,i) #รายได้สุทธิ
                    filename = f"{self.get_value(sheet,2,i)},{self.get_value(sheet,21,i)},0,{date_m.strftime('%B')},{datetime.now().strftime('%d%m%y%H%M%S')}"
                    finalpath_ex = os.path.join(self.output_dir,sheet.title,f"{filename}.xlsx")
                    self.temporary.save(finalpath_ex)
                    file.append(salib['C6'].value)
                    time.sleep(0.3)
                    wb = app.Workbooks.Open(finalpath_ex)
                    wb.ActiveSheet.PageSetup.Orientation = 2
                    wb.ActiveSheet.PageSetup.Zoom = False
                    wb.ActiveSheet.PageSetup.FitToPagesTall = 1
                    wb.ActiveSheet.PageSetup.FitToPagesWide = 1
                    wb.ActiveSheet.ExportAsFixedFormat(0,os.path.join(self.output_dir,sheet.title,f"{filename}"))
                    wb.Save()
                    wb.Close()   
                    time.sleep(0.1)
                    os.remove(finalpath_ex)
            shutil.copyfile(self.path,self.temporaries)
            self.temporary = load_workbook(self.temporaries,data_only=True)
        os.remove(self.temporaries)
        self.complete = True
        self.progress()
        creating = False

class Content(BoxLayout):
    pass

class list_container(IRightBodyTouch, MDBoxLayout):
    adaptive_height=True

class SlipMaker(MDScreen):
    def on_start(self):
        pass

    excel_path = StringProperty("Please select file")
    going_to_make_slip = []
    excel_object:excel = None
    total_individual_checkbox = 0

    def call_back_create_slip(self,data):
        print(data)
        if data['status'] == 'processing':
            Clock.schedule_once(lambda dt: self.update_progress_slip_bar(data))
        elif data['status'] == 'complete':   
            Clock.schedule_once(lambda dt: self.done_create_slip(),1)

    def done_create_slip(self):
        self.progress_dialog.title = "Complete !!"
        self.progress_dialog.dismiss()
        snackbar = MDSnackbar(
                MDLabel(
                text="Dowload Complete !",
                font_name = 'sarabun',
                text_color=self.theme_cls.opposite_bg_normal,
            ),
            pos_hint={"center_x": 0.5},
            size_hint_x=0.5,
            md_bg_color="#ffffff",
            duration = 5
            )
        snackbar.open()

    def update_progress_slip_bar(self, data):
        self.ids['progress_slip_bar'].value = data['percentage']
        self.ids['label_slip'].text = f"{int(data['percentage'])}% creating file for {data['branch']}/{data['current']}"

    def start_creating_slip(self, instance, value:date, date_range):
        self.progress_dialog = MDDialog(
            title=f'[size=28][font=sarabunBold]Creating files[/font][/size]',
            type="custom",
            content_cls=MDBoxLayout(
            orientation="vertical",
            spacing= "30dp",
            padding= "20dp",
            size_hint_y = None,
            height= "50dp",
            ),
            radius=[20, 7, 20, 7]
        )
        self.progress_dialog.auto_dismiss = False
        progress_bar = MDProgressBar(value=0,pos_hint={'center_y':-0.2})
        label = MDLabel(text='Initualizing',theme_text_color='Hint',pos_hint={'center_y':0},font_style='sarabun',font_size= "20dp")
        self.ids['label_slip'] = label
        self.ids['progress_slip_bar'] = progress_bar
        self.progress_dialog.content_cls.add_widget(label)
        self.progress_dialog.content_cls.add_widget(progress_bar)
        self.progress_dialog.open()
        self.excel_object.call_back = self.call_back_create_slip
        thread = threading.Thread(target=self.excel_object.extract_convert, args=(self.going_to_make_slip,value,))
        thread.start()

    def close_confirm_makeslip_dialog(self,dialog):
        self.confirm_makeslip_dialog.dismiss()

    def select_date(self,dialog):
        self.confirm_makeslip_dialog.dismiss()
        today = datetime.today()
        self.date_dialog = MDDatePicker(
            year=today.year+543,
            month=today.month,
            day=today.day,
            min_year=2400,
            max_year=2666,
            title_input="Input payment date",
            title="Select payment date",
            )
        self.date_dialog.bind(on_save=self.start_creating_slip)
        self.date_dialog.open()

    def create_slip(self,button):
        cancle_button = MDFlatButton(
                        text="CANCEL",
                        on_release=self.close_confirm_makeslip_dialog
                    )
        start_button = MDRaisedButton(
                        text="Next",
                        on_release=self.select_date
                    )
        self.confirm_makeslip_dialog = MDDialog(
            title=f'[size=28][font=sarabunBold]Are you sure to start making slip for {len(self.going_to_make_slip)} people?[/font][/size]',
            text='[size=20][font=sarabun]This process can not be canceled make sure you are ready[/font][/size]',
            type="simple",
            radius=[20, 7, 20, 7],
            buttons=[cancle_button,start_button]
        )
        print(self.going_to_make_slip)
        self.confirm_makeslip_dialog.open()
    
    def individual_selected(self,button:MDCheckbox,pos):
        text = button.parent.parent.parent.text
        pattern = r'\[.*?\]\d+.(.*?)\[/.*?]'
        result = re.search(pattern, text)
        name = result.group(1)
        if button.active:
            self.going_to_make_slip.append(name)
        else:
            try:
                self.going_to_make_slip.remove(name)
            except:pass

        true_state = len(self.going_to_make_slip)
        self.ids['number'].text = str(true_state)
        if true_state > 0:
            self.ids.create_slip.disabled=False
        else:
            self.ids.create_slip.disabled=True

        if true_state == self.total_individual_checkbox:
            self.ids.all_select.active = True
        else:
            self.ids.all_select.active = False


    def all_selected(self,button:MDRaisedButton):
        state = button.active
        for ids in self.ids: 
            if 'checkbox' in ids:
                self.ids[ids].active = state

    def create_employee_list(self,dt):
        fram1 = MDBoxLayout(
            orientation="vertical",
            md_bg_color= (0.217,0.217,0.217,0.1),
            size_hint= (0.5,1),
            id = 'fram1'
        )
        self.ids['fram1'] = fram1
        fram1.add_widget(
            MDLabel(
                text= "Employee List",
                size_hint= (1,0.1),
                halign= 'center',
                font_style= "sarabunBold",
                font_size= "20sp"
            )
        )
        fram2 = MDBoxLayout(
            orientation="horizontal",
            size_hint= (1,0.08)
        )
        fram2.add_widget(
            MDLabel(
                text= "All",
                size_hint= (0.3,1),
                pos_hint={'center_y':0.75},
                halign= 'right',
                font_style= "sarabun",
                font_size= "25dp"
            )
        )
        all_checkbox = MDCheckbox(
                size_hint= (0.3,1),
                pos_hint={'center_y':0.65},
                halign= 'left',
            ) 
        self.ids["all_select"] = all_checkbox
        all_checkbox.bind(on_press=self.all_selected)
        fram2.add_widget(all_checkbox)
        fram1.add_widget(fram2)
        fram1.add_widget(MDSeparator())
        scroll = MDScrollView()  
        mdlst_slip = MDList()
        self.ids['lst'] = mdlst_slip
        for branch in self.excel_object.sources:
            amount = self.excel_object.get_round(branch)
            content = MDBoxLayout()
            content.adaptive_height = True
            content.orientation = 'vertical'
            self.ids[branch.title+'slip'] = content 
            panel = MDExpansionPanel(
                    icon="office-building-outline",
                    content=content,
                    panel_cls=MDExpansionPanelTwoLine(
                        text=f"[size=25]{branch.title}[/size]",
                        secondary_text=f"{amount} คน",
                        font_style= "sarabunBold",
                        secondary_font_style= "sarabunBold"
                    )
                )
            mdlst_slip.add_widget(panel)
            for num,rows in enumerate(range(amount),1):
                rows += 3
                item = OneLineAvatarIconListItem(
                    text=f"[size=20]{num}.{self.excel_object.get_value(branch,2,rows)}[/size]",
                    font_style = "sarabun",
                )
                face = IconLeftWidget(
                    icon='account'
                )
                check = list_container()
                checkbox = MDCheckbox()
                checkbox.bind(active=self.individual_selected)
                self.total_individual_checkbox += 1
                self.ids[f'checkbox{self.excel_object.get_value(branch,3,rows)}'] = checkbox
                checkbox.color_active = self.theme_cls.accent_light
                item.add_widget(face)
                check.add_widget(checkbox)
                item.add_widget(check)
                b = str(branch.title) +'slip'
                self.ids[b].add_widget(item)
        scroll.add_widget(mdlst_slip)
        fram1.add_widget(scroll)
        fram1.add_widget(MDSeparator())
        screen = MDScreen(
            size_hint = (1,0.2)
        )
        create_button = MDRaisedButton(
                size_hint= (None,None),
                text= "Create Slip",
                pos_hint= {'center_x':.6,'center_y':.5},
                disabled = BooleanProperty(True)
            )
        self.ids['create_slip'] = create_button
        create_button.bind(on_press=self.create_slip)
        ico = MDIcon(
            icon='account',
            pos_hint= {'center_x':.1,'center_y':.53},
        )
        number = MDLabel(
            text = "0",
            font_name='sarabun',
            pos_hint= {'center_x':0.7,'center_y':.53}
        )
        self.ids['number'] = number
        screen.add_widget(number)
        screen.add_widget(ico)
        screen.add_widget(create_button)
        fram1.add_widget(screen)
        self.ids.box1.add_widget(fram1)

    def on_browse(self):
        path = self.file_open()
        if path:
            self.excel_object = excel(path = path)
            try:
                self.ids.box1.remove_widget(self.ids.fram1)
            except:pass
            Clock.schedule_once(self.create_employee_list)
            self.snackbar = MDSnackbar(
                MDLabel(
                text="Upload Complete !",
                font_name = 'sarabun',
                text_color=self.theme_cls.opposite_bg_normal
            ),
            pos_hint={"center_x": 0.5},
            size_hint_x=0.5,
            md_bg_color="#ffffff"
            )
            self.snackbar.open()

    def file_open(self):
        root = tkinter.Tk()
        root.withdraw()
        # root.iconbitmap("YOUR_IMAGE.ico")
        currdir = os.getcwd()
        tempdir = filedialog.askopenfilename(parent=root, initialdir=currdir, title='Please select excel file',filetypes=[("Excel files",".xlsx .xls")])
        path = None
        if len(tempdir) > 0:
            path = tempdir
            self.excel_path = os.path.basename(tempdir)
        return path
    
class CustomOneLineAvatarIconListItem(OneLineAvatarIconListItem):
    def __init__(self, path,email,name,branch, file_name,*args, **kwargs):
        super().__init__(*args, **kwargs)
        self.path = path
        self.name = name
        self.email = email
        self.branch = branch
        self.file_name = file_name

class GmailSender(MDScreen):
    going_to_send = []
    recent_employee = 0
    total_individual_checkbox = 0
    def on_start(self):
        Clock.schedule_interval(self.update_lst,0.5)

    def done_send_email(self):
        self.ids['box_email'].clear_widgets()
        Clock.schedule_once(lambda dt: self.add_lst())
        self.progress_dialog.title = "Complete !!"
        self.progress_dialog.dismiss()
        snackbar = MDSnackbar(
                MDLabel(
                text="Send Complete !",
                font_name = 'sarabun',
                text_color=self.theme_cls.opposite_bg_normal,
            ),
            pos_hint={"center_x": 0.5},
            size_hint_x=0.5,
            md_bg_color="#ffffff",
            duration = 5
            )
        snackbar.open()

    def call_back_create_email(self,data):
        print(data)
        if data['status'] == 'processing':
            Clock.schedule_once(lambda dt: self.update_progress_email_bar(data))
        elif data['status'] == 'complete':   
            Clock.schedule_once(lambda dt: self.done_send_email(),1)

    def update_progress_email_bar(self, data):
        self.ids['progress_email_bar'].value = data['percentage']
        self.ids['label_email'].text = f"{int(data['percentage'])}% Sending file for {data['branch']}/{data['current']}"

    def start_send_email(self,dialog):
        self.confirm_sendemail_dialog.dismiss()
        # for path in os.listdir('slip'):
        #     for file in os.listdir('slip/'+path):
        #         name = file.split(',')
        #         name[-1] = '1.pdf'
        #         new_name = ','.join(name)
        #         print(self.going_to_send)
        #         print('----------')
        #         print('slip/'+path+'/'+file)
        #         if 'slip/'+path+'/'+file in self.going_to_send:
        #             os.rename(f'slip/{path}/{file}',f'slip/{path}/{new_name}')
        #             self.ids['box_email'].clear_widgets()
        #             Clock.schedule_once(lambda dt: self.add_lst())
        self.progress_dialog = MDDialog(
            title=f'[size=28][font=sarabunBold]Sending mail[/font][/size]',
            type="custom",
            content_cls=MDBoxLayout(
            orientation="vertical",
            spacing= "30dp",
            padding= "20dp",
            size_hint_y = None,
            height= "50dp"
            ),
            radius=[20, 7, 20, 7]
        )
        self.progress_dialog.auto_dismiss = False
        progress_bar = MDProgressBar(value=0,pos_hint={'center_y':-0.2})
        label = MDLabel(text='Initualizing',theme_text_color='Hint',pos_hint={'center_y':0},font_style='sarabun',font_size= "20dp")
        self.ids['label_email'] = label
        self.ids['progress_email_bar'] = progress_bar
        self.progress_dialog.content_cls.add_widget(label)
        self.progress_dialog.content_cls.add_widget(progress_bar)
        self.progress_dialog.open()
        self.sendemail_object = send_email()
        self.sendemail_object.call_back = self.call_back_create_email
        thread = threading.Thread(target=self.sendemail_object.send, args=(self.going_to_send,))
        thread.start()

    def close_confirm_sendemail_dialog(self,dialog):
        self.confirm_sendemail_dialog.dismiss()

    def send_email_button(self):
        cancle_button = MDFlatButton(
                        text="CANCEL",
                        on_release=self.close_confirm_sendemail_dialog
                    )
        start_button = MDRaisedButton(
                        text="Start",
                        on_release=self.start_send_email
                    )
        self.confirm_sendemail_dialog = MDDialog(
            title=f'[size=28][font=sarabunBold]Are you sure to start send email for {len(self.going_to_send)} people?[/font][/size]',
            text='[size=20][font=sarabun]This process may take a several minutes and can not be cancled[/font][/size]',
            radius=[20, 7, 20, 7],
            buttons=[cancle_button,start_button]
        )
        self.confirm_sendemail_dialog.open()

    def update_lst(self,dt):
        f = []
        for path in os.listdir('slip'):
            for i in os.listdir('slip/'+path):
                f.append(i)
        if creating == True:
                return
        if len(f) != self.recent_employee:
            self.ids['box_email'].clear_widgets()
            if len(f) != 0:
                Clock.schedule_once(lambda dt: self.add_lst())
                
            else:
                Clock.schedule_once(lambda dt: self.no_employee_label())
                self.ids.sendemail_button.disabled=True
                self.ids.sendemail_button.text = f' 0       Send email               '
            self.recent_employee = len(f)

    def individual_selected(self,button:MDCheckbox,pos):
        data = button.parent.parent.parent
        payload={
            'path':data.path,
            'name':data.name,
            'email':data.email,
            'branch':data.branch,
            'file_name':data.file_name
        }
        if button.active:
            self.going_to_send.append(payload)
        else:
            try:
                self.going_to_send.remove(payload)
            except:pass

        true_state = len(self.going_to_send)
        self.ids.sendemail_button.text = f' {true_state}       Send email               '

        if true_state > 0:
            self.ids.sendemail_button.disabled=False
        else:
            self.ids.sendemail_button.disabled=True

        if true_state == self.total_individual_checkbox:
            self.ids.all_select_email.active = True
        else:
            self.ids.all_select_email.active = False


    def all_selected(self,button:MDRaisedButton):
        state = button.active
        for ids in self.ids: 
            if 'checkbox_email_' in ids:
                self.ids[ids].active = state

    def add_lst(self):
        self.going_to_send=[]
        self.ids.sendemail_button.text = f' {len(self.going_to_send)}       Send email               '
        fram1 = MDBoxLayout(
            orientation="horizontal",
            size_hint= (1,0.08)
        )
        self.ids['fram1_employee'] = fram1
        fram1.add_widget(
            MDLabel(
                text= "All",
                size_hint= (0.3,1),
                pos_hint={'center_y':0.75},
                halign= 'right',
                font_style= "sarabun",
                font_size= "25dp"
            )
        )
        all_checkbox = MDCheckbox(
            size_hint= (0.3,1),
            pos_hint={'center_y':0.65},
            halign= 'left',
        ) 
        all_checkbox.bind(on_press=self.all_selected)
        self.ids["all_select_email"] = all_checkbox
        fram1.add_widget(all_checkbox)
        self.ids['box_email'].add_widget(fram1)
        scroll = MDScrollView()
        self.ids['scroll_email'] = scroll
        mdlst_employee = MDList()
        scroll.add_widget(mdlst_employee)

        for branch in os.listdir('slip'):
            content = MDBoxLayout()
            content.adaptive_height = True
            content.orientation = 'vertical'
            self.ids[f"{branch.title}email"] = content 
            panel = MDExpansionPanel(
                icon="office-building-outline",
                content=content,
                pos_hint={'center_y':1},
                panel_cls=MDExpansionPanelTwoLine(
                    text=f"[size=25]{branch}[/size]",
                    secondary_text=f"{len(os.listdir('slip/'+branch))}คน",
                    font_style= "sarabunBold",
                    secondary_font_style= "sarabunBold"
                )
            )
            #39
            for num,filename in enumerate(os.listdir('slip/'+branch),1):
                # cleaned_string = regex.sub('\p{M}', '',f"{num}.{filename.split(',')[0]}")
                # name_lenght = len(cleaned_string)
                # less = 39 - name_lenght
                # space = ' '*less)
                name,email,checker,ofmonth,createat = filename[:-4].split(',')
                print(createat)
                item = CustomOneLineAvatarIconListItem(
                    text=f"[size=20]{num}.{name}   |   {email}   |   Salary of [b]{ofmonth}[/b]   [color=9C9B9B]Create at {datetime.strptime(createat,'%d%m%y%H%M%S').strftime('%d/%m/%y %H:%M')}[/color][/size]",
                    font_style = "sarabun",
                    path= os.path.join('slip',branch,filename),
                    email=filename.split(',')[1],
                    name=filename.split(',')[0],
                    branch=branch,
                    file_name=filename
                )
                if checker == '0':
                    icon = 'account-alert'
                else:
                    icon = 'account-check'
                face = IconLeftWidget(
                    icon=icon
                )
                check = list_container()
                checkbox = MDCheckbox()
                checkbox.bind(active=self.individual_selected)
                self.total_individual_checkbox += 1
                self.ids[f'checkbox_email_{filename.split(",")[0]}'] = checkbox
                checkbox.color_active = self.theme_cls.accent_light
                item.add_widget(face)
                check.add_widget(checkbox)
                item.add_widget(check)
                b = str(branch.title)+'email'
                self.ids[b].add_widget(item)
            mdlst_employee.add_widget(panel)
        self.ids['box_email'].add_widget(scroll)

    def no_employee_label(self):
        for path in os.listdir('slip'):
            try:
                os.rmdir(f'slip/{path}')
            except:pass
        label = MDLabel(
                text= 'No employee deploy',
                halign= 'center',
                pos_hint={"center_y":.5},
                theme_text_color='Hint')
        self.ids['no_employee_label_email'] = label
        self.ids['box_email'].add_widget(label)

class Employee(MDScreen):
    recent_employee = 0
    total_individual_checkbox = 0
    going_to_delete = []

    def on_start(self):
        Clock.schedule_interval(self.update_lst,0.5)

    def update_lst(self,dt):
        f = []
        for path in os.listdir('slip'):
            for i in os.listdir('slip/'+path):
                f.append(i)
        if creating == True:
            return
        if len(f) != self.recent_employee:
            self.ids['box_employee'].clear_widgets()
            if len(f) != 0:
                Clock.schedule_once(lambda dt: self.add_lst())
            else:
                Clock.schedule_once(lambda dt: self.no_employee_label())
                self.ids.delete_button.disabled=True
                self.ids.delete_label_count.text = '0'
            self.recent_employee = len(f)

    def deletion(self, path, file):
        try:
            os.remove(f'slip/{path}/{file}')
        except Exception as e:
            print(e)

    def delete_selected(self):
        for path in os.listdir('slip'):
            for file in os.listdir('slip/'+path):
                name = file.split(',')[0]
                if name in self.going_to_delete:
                    thread = threading.Thread(target=self.deletion, args=(path,file,))
                    thread.start()
        print(self.going_to_delete)            
        snackbar = MDSnackbar(
                MDLabel(
                text="Deleted !",
                font_name = 'sarabun',
                text_color=self.theme_cls.opposite_bg_normal,
            ),
            pos_hint={"center_x": 0.5},
            size_hint_x=0.5,
            md_bg_color="#ffffff",
            duration = 3
            )
        snackbar.open()

    def individual_selected(self,button:MDCheckbox,pos):
        text = button.parent.parent.parent.text
        pattern = r'\[.*?\]\d+.(.*?)\[/.*?]'
        result = re.search(pattern, text)
        name = result.group(1)
        if button.active:
            self.going_to_delete.append(name)
        else:
            try:
                self.going_to_delete.remove(name)
            except:pass

        true_state = len(self.going_to_delete)

        self.ids.delete_label_count.text = str(true_state)

        if true_state > 0:
            self.ids.delete_button.disabled=False
        else:
            self.ids.delete_button.disabled=True

        if true_state == self.total_individual_checkbox:
            self.ids.all_select_employee.active = True
        else:
            self.ids.all_select_employee.active = False


    def all_selected(self,button:MDRaisedButton):
        state = button.active
        for ids in self.ids: 
            if 'checkbox_employee_' in ids:
                self.ids[ids].active = state
                
    
    def add_lst(self):
        self.going_to_delete=[]
        self.ids.delete_label_count.text = str(len(self.going_to_delete))
        fram1 = MDBoxLayout(
            orientation="horizontal",
            size_hint= (1,0.08)
        )
        self.ids['fram1_employee'] = fram1
        fram1.add_widget(
            MDLabel(
                text= "All",
                size_hint= (0.3,1),
                pos_hint={'center_y':0.75},
                halign= 'right',
                font_style= "sarabun",
                font_size= "25dp"
            )
        )
        all_checkbox = MDCheckbox(
            size_hint= (0.3,1),
            pos_hint={'center_y':0.65},
            halign= 'left',
        ) 
        all_checkbox.bind(on_press=self.all_selected)
        self.ids["all_select_employee"] = all_checkbox
        fram1.add_widget(all_checkbox)
        self.ids['box_employee'].add_widget(fram1)
        scroll = MDScrollView()
        self.ids['scroll_employee'] = scroll
        mdlst_employee = MDList()
        scroll.add_widget(mdlst_employee)

        for branch in os.listdir('slip'):
            content = MDBoxLayout()
            content.adaptive_height = True
            content.orientation = 'vertical'
            self.ids[f"{branch.title}employee"] = content 
            panel = MDExpansionPanel(
                icon="office-building-outline",
                content=content,
                pos_hint={'center_y':1},
                panel_cls=MDExpansionPanelTwoLine(
                    text=f"[size=25]{branch}[/size]",
                    secondary_text=f"{len(os.listdir('slip/'+branch))}คน",
                    font_style= "sarabunBold",
                    secondary_font_style= "sarabunBold"
                )
            )
            for num,filename in enumerate(os.listdir('slip/'+branch),1):
                item = OneLineAvatarIconListItem(
                    text=f"[size=20]{num}.{filename.split(',')[0]}[/size]",
                    font_style = "sarabun",
                )
                face = IconLeftWidget(
                    icon='account'
                )
                check = list_container()
                checkbox = MDCheckbox()
                checkbox.bind(active=self.individual_selected)
                self.total_individual_checkbox += 1
                self.ids[f'checkbox_employee_{filename.split(",")[0]}'] = checkbox
                checkbox.color_active = self.theme_cls.accent_light
                item.add_widget(face)
                check.add_widget(checkbox)
                item.add_widget(check)
                b = str(branch.title)+'employee'
                self.ids[b].add_widget(item)
            mdlst_employee.add_widget(panel)
        self.ids['box_employee'].add_widget(scroll)

    def no_employee_label(self):
        for path in os.listdir('slip'):
            try:
                os.rmdir(f'slip/{path}')
            except:pass
        label = MDLabel(
                text= 'No employee deploy',
                halign= 'center',
                pos_hint={"center_y":.5},
                theme_text_color='Hint')
        self.ids['no_employee_label'] = label
        self.ids['box_employee'].add_widget(label)

class Setting(MDScreen):
    def on_start(self):
        pass

class SliperApp(MDApp):
    employee = []
    def build(self):
        LabelBase.register(name='sarabun',
            fn_regular=r'data\font\THSarabunNew.ttf',
        )
        theme_font_styles.append('sarabun')
        self.theme_cls.font_styles["sarabun"] = [
            "sarabun",
            20,
            False,
            0.15,
        ]
        LabelBase.register(name='sarabunBold',
            fn_regular=r'data\font\THSarabunNew Bold.ttf'
        )
        theme_font_styles.append('sarabunBold')
        self.theme_cls.font_styles["sarabunBold"] = [
            "sarabunBold",
            20,
            False,
            0.15,
        ]
        self.theme_cls.primary_palette = "Brown"
        self.theme_cls.primary_hue = "400"
        self.theme_cls.theme_style = "Light"
        screen = Builder.load_string(KV)
        return screen
    
    def on_start(self):
       sc_lst = self.root.ids.mana.screen_names
       for screen in sc_lst:
           self.root.ids.mana.get_screen(screen).on_start()
    
    
    


SliperApp().run()